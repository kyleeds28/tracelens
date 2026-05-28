from pathlib import Path
from typing import Any
import yaml
import openpyxl


def load_mapping(mapping_path: Path) -> dict:
    with mapping_path.open("r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def _norm_kind(raw: Any, kind_values: dict) -> str | None:
    if raw is None:
        return None
    s = str(raw).strip()
    for k, vals in kind_values.items():
        if s in vals:
            return k
    return None


def _norm_label(s: Any) -> str:
    """Whitespace-stripped, lowercased label for tolerant matching."""
    if s is None:
        return ""
    return "".join(str(s).split()).lower()


def _resolve_sheet(wb, want: str, aliases: list[str], col_map: dict[str, str],
                   header_row: int) -> str:
    """Pick the right sheet from `wb` using progressively looser rules.

    Order:
      1. exact name (or any alias)
      2. whitespace-insensitive equality
      3. substring (normalized) either direction
      4. auto-discovery: scan every sheet, pick the first whose header_row
         contains all the configured column labels
    """
    names = wb.sheetnames
    candidates = [want] + [a for a in (aliases or []) if a and a != want]

    # 1) exact
    for c in candidates:
        if c in names:
            return c

    # 2) whitespace-insensitive equality
    norm_names = {_norm_label(n): n for n in names}
    for c in candidates:
        hit = norm_names.get(_norm_label(c))
        if hit:
            return hit

    # 3) substring (normalized) either direction
    for c in candidates:
        nc = _norm_label(c)
        for nn, original in norm_names.items():
            if nc and (nc in nn or nn in nc):
                return original

    # 4) auto-discover by header columns presence
    if col_map:
        wanted_headers = {_norm_label(h) for h in col_map.values()}
        for n in names:
            try:
                ws = wb[n]
                row_vals = {
                    _norm_label(ws.cell(row=header_row, column=c).value)
                    for c in range(1, ws.max_column + 1)
                }
                # accept if at least 60% of required columns are present
                hit = wanted_headers & row_vals
                if len(hit) >= max(1, int(len(wanted_headers) * 0.6)):
                    return n
            except Exception:
                continue

    raise ValueError(
        f"프로그램 목록 시트를 찾지 못했습니다. "
        f"기대한 이름: {candidates}. xlsx 시트들: {names}. "
        f"기대한 헤더 컬럼들: {list(col_map.values())}"
    )


def _options(value) -> list[str]:
    """Coerce a column-mapping entry to a list of acceptable header strings."""
    if value is None:
        return []
    if isinstance(value, list):
        return [str(v) for v in value if v is not None]
    return [str(value)]


def _detect_header_row(ws, col_map: dict, candidates: list[int]) -> int:
    """Among `candidates`, pick the row whose header labels match the most
    configured columns (tie-break: smallest row number)."""
    if not candidates:
        return 1
    best_row = candidates[0]
    best_score = -1
    for r in candidates:
        try:
            row_set = {
                _norm_label(ws.cell(row=r, column=c).value)
                for c in range(1, ws.max_column + 1)
                if ws.cell(row=r, column=c).value is not None
            }
        except Exception:
            continue
        score = 0
        for opts in col_map.values():
            for o in _options(opts):
                if _norm_label(o) in row_set:
                    score += 1
                    break
        if score > best_score:
            best_score = score
            best_row = r
    return best_row


def _resolve_columns(ws, header_row: int, col_map: dict) -> tuple[dict[str, int], list[str]]:
    """Return ({norm_name: col_idx}, missing_required_names).

    A column entry that is a *list* tries each alternative until one matches.
    A column entry that is a *string* is a single required label.
    Optional columns (none of the alternatives present) are simply omitted
    from the return map — callers can `.get(...)` safely later.
    """
    header_cells = {
        _norm_label(ws.cell(row=header_row, column=c).value): c
        for c in range(1, ws.max_column + 1)
        if ws.cell(row=header_row, column=c).value is not None
    }
    name_to_col: dict[str, int] = {}
    missing_required: list[str] = []
    REQUIRED = {"module_name"}   # the only column truly required to do any matching
    for norm_name, opts in col_map.items():
        col = None
        for label in _options(opts):
            col = header_cells.get(_norm_label(label))
            if col is not None:
                break
        if col is not None:
            name_to_col[norm_name] = col
        elif norm_name in REQUIRED:
            missing_required.append(norm_name)
    return name_to_col, missing_required


def load_spec(xlsx_path: Path, mapping_path: Path) -> list[dict]:
    cfg = load_mapping(mapping_path)
    spec = cfg["spec"]
    sheet_name = spec["sheet"]
    sheet_aliases: list[str] = spec.get("sheet_aliases", []) or []
    # accept either a single header_row OR a list of candidates
    header_row_candidates: list[int] = (
        [int(r) for r in spec["header_row_candidates"]]
        if "header_row_candidates" in spec
        else [int(spec.get("header_row", 1))]
    )
    col_map: dict = spec["columns"]
    forward_fill: list[str] = spec.get("forward_fill", [])
    kind_values: dict[str, list[str]] = spec.get("kind_values", {})

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # Sheet resolution still uses one header_row for its auto-discovery; we feed
    # the first candidate. The actual header row used below is detected per ws.
    sheet_name = _resolve_sheet(wb, sheet_name, sheet_aliases, col_map, header_row_candidates[0])
    ws = wb[sheet_name]

    header_row = _detect_header_row(ws, col_map, header_row_candidates)
    name_to_col, missing_required = _resolve_columns(ws, header_row, col_map)

    if missing_required:
        # collect headers from the chosen row for the error message
        seen = [
            ws.cell(row=header_row, column=c).value
            for c in range(1, ws.max_column + 1)
            if ws.cell(row=header_row, column=c).value is not None
        ]
        wb.close()
        raise ValueError(
            f"필수 컬럼을 찾지 못했습니다: {missing_required}. "
            f"시트 '{sheet_name}' 의 헤더(row {header_row}): {seen}"
        )

    rows: list[dict] = []
    last: dict[str, Any] = {}

    for r in range(header_row + 1, ws.max_row + 1):
        row: dict[str, Any] = {"row_idx": r}
        all_none = True
        for norm_name, col_idx in name_to_col.items():
            v = ws.cell(row=r, column=col_idx).value
            if v is not None:
                all_none = False
            row[norm_name] = v

        if all_none:
            continue

        # forward fill
        for ff in forward_fill:
            if row.get(ff) in (None, ""):
                row[ff] = last.get(ff)
            else:
                last[ff] = row[ff]
        # also remember last for non-ff keys (no fill, but track)
        for k in name_to_col:
            if k not in forward_fill and row.get(k) not in (None, ""):
                last[k] = row[k]

        row["kind_norm"] = _norm_kind(row.get("kind"), kind_values)
        rows.append(row)

    wb.close()
    return rows


def group_by_program(rows: list[dict]) -> list[dict]:
    """Group flat rows by program_id; produce one program object per group."""
    by_id: dict[str, dict] = {}
    order: list[str] = []
    for r in rows:
        pid = r.get("program_id") or f"_anon_{r['row_idx']}"
        if pid not in by_id:
            by_id[pid] = {
                "program_id": pid,
                "program_name": r.get("program_name"),
                "category_l1": r.get("category_l1"),
                "category_l2": r.get("category_l2"),
                "menu_url": r.get("menu_url"),
                "rows": [],
            }
            order.append(pid)
        g = by_id[pid]
        # prefer the first non-null program_name we encounter for the group
        if not g["program_name"] and r.get("program_name"):
            g["program_name"] = r["program_name"]
        if not g["menu_url"] and r.get("menu_url"):
            g["menu_url"] = r["menu_url"]
        g["rows"].append(r)
    return [by_id[pid] for pid in order]
